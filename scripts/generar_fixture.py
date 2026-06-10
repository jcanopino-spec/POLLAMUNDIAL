# Genera public/fixture-mundial.xlsx: fixture COMPLETO del Mundial 2026 con
# tablas de grupos AUTO-CALCULADAS por fórmulas (mete marcadores → se arman solas).
# Uso: python3 scripts/generar_fixture.py
import json
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INK="1B1714"; CREAM="FBF4E6"; CREAM2="F3E9D2"; YELLOW="FFC22E"; GREEN="1B9150"; RED="E1382F"; BLUE="3447D6"; PAPER="FFFFFF"

ES = {
 "Mexico":"México","South Africa":"Sudáfrica","Korea Republic":"Corea del Sur","Czechia":"Chequia",
 "Canada":"Canadá","Bosnia and Herzegovina":"Bosnia","Qatar":"Catar","Switzerland":"Suiza",
 "Brazil":"Brasil","Morocco":"Marruecos","Haiti":"Haití","Scotland":"Escocia",
 "USA":"Estados Unidos","Paraguay":"Paraguay","Australia":"Australia","Türkiye":"Turquía",
 "Germany":"Alemania","Curaçao":"Curazao","Côte d'Ivoire":"Costa de Marfil","Ecuador":"Ecuador",
 "Netherlands":"Países Bajos","Japan":"Japón","Sweden":"Suecia","Tunisia":"Túnez",
 "Belgium":"Bélgica","Egypt":"Egipto","IR Iran":"Irán","New Zealand":"Nueva Zelanda",
 "Spain":"España","Cabo Verde":"Cabo Verde","Saudi Arabia":"Arabia Saudita","Uruguay":"Uruguay",
 "France":"Francia","Senegal":"Senegal","Norway":"Noruega","Iraq":"Irak",
 "Argentina":"Argentina","Algeria":"Argelia","Austria":"Austria","Jordan":"Jordania",
 "Colombia":"Colombia","Portugal":"Portugal","Uzbekistan":"Uzbekistán","Congo DR":"RD Congo",
 "England":"Inglaterra","Croatia":"Croacia","Ghana":"Ghana","Panama":"Panamá",
}
FASE={1:"Grupos F1",2:"Grupos F2",3:"Grupos F3",4:"16avos",5:"Octavos",6:"Cuartos",7:"Semis",8:"Final/3º"}
DIAS=["lun","mar","mié","jue","vie","sáb","dom"]; MESES=["ene","feb","mar","abr","may","jun","jul","ago","sep","oct","nov","dic"]
BOG=ZoneInfo("America/Bogota")

def nombre(eq):
    if eq in ES: return ES[eq]
    import re
    m=re.match(r"^([123])([A-L])$",eq);  m2=re.match(r"^3([A-L]{2,})$",eq); m3=re.match(r"^W(\d+)$",eq); m4=re.match(r"^(?:L|RU)(\d+)$",eq)
    if m: return f"{m.group(1)}º Gpo {m.group(2)}"
    if m2: return "3º de "+"/".join(m2.group(1))
    if m3: return f"Ganador P{m3.group(1)}"
    if m4: return f"Perdedor P{m4.group(1)}"
    return eq

def fecha(m):
    dt=datetime.fromisoformat(m["DateUtc"].replace(" ","T").replace("Z","+00:00")).astimezone(BOG)
    return f"{DIAS[dt.weekday()]} {dt.day} {MESES[dt.month-1]} · {dt.strftime('%I:%M %p').lstrip('0')}"

fixture=sorted(json.load(open("data/fixture-raw.json")),key=lambda m:(m["DateUtc"],m["MatchNumber"]))
thin=Side(style="thin",color=INK); med=Side(style="medium",color=INK)
B=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=Workbook()

def style(c,fnt=None,fill=None,align="center",border=True,wrap=False):
    if fnt: c.font=fnt
    if fill: c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if border: c.border=B

H=Font(bold=True,color=PAPER,size=10); N=Font(size=10); Bd=Font(bold=True,size=10)
TIT=Font(name="Arial Black",size=15,color=INK)

# ===================== HOJA 1: LÉEME =====================
ws=wb.active; ws.title="LÉEME 🐔"; ws.sheet_properties.tabColor=YELLOW
ws.column_dimensions["A"].width=104
filas=[
 ("🐔⚽ FIXTURE MUNDIAL 2026 — LA POLLA DE ALAMEDA (versión Excel) ⚽🐔",TIT,YELLOW),
 ("",None,None),
 ("Pa'l parcero al que le gusta el Excel más que el WhatsApp 😂. Este archivo hace la magia solo:",N,None),
 ("",None,None),
 ("📋 HOJA 'FASE DE GRUPOS': escribe los goles en las casillas amarillas y…",Bd,CREAM),
 ("📊 HOJA 'TABLAS': las posiciones de los 12 grupos se calculan SOLAS (puntos, goles, diferencia y posición). ¡No tocas nada!",N,None),
 ("🏆 HOJA 'ELIMINATORIAS': el cuadro completo desde 16avos hasta la final, con fechas y estadios (hora Colombia).",N,None),
 ("",None,None),
 ("⚽ Recuerda: ganar = 3 pts, empate = 1, perder = 0. Clasifican los 2 primeros de cada grupo + los 8 mejores terceros.",N,None),
 ("📈 Si dos quedan iguales: desempata la diferencia de goles, luego los goles a favor.",N,None),
 ("",None,None),
 ("🎯 ¿Quieres JUGAR la polla de verdad y ganar billete pa' la natillera? Esto es solo pa' practicar.",Bd,None),
 ("   La buena está en la app (te calcula puntos, te avisa morosos y trae hasta simulador): pollamundialnatillera.vercel.app",Font(bold=True,color=BLUE,size=11),None),
 ("",None,None),
 ("Hecho con cariño por INPLUX SAS · prohibido copiarlo: la gallina tiene abogados 🐔⚖️",Font(italic=True,size=9),None),
]
for i,(t,f,fl) in enumerate(filas,1):
    c=ws.cell(row=i,column=1,value=t)
    if f: c.font=f
    if fl: c.fill=PatternFill("solid",fgColor=fl)
    c.alignment=Alignment(wrap_text=True,vertical="center")

# ===================== HOJA 2: FASE DE GRUPOS (editable, ordenada por grupo) =====================
wg=wb.create_sheet("FASE DE GRUPOS"); wg.sheet_properties.tabColor=GREEN
grupos={}
for m in fixture:
    if m["Group"]:
        g=m["Group"].replace("Group ","")
        grupos.setdefault(g,[]).append(m)
heads=["P#","Grupo","LOCAL","GL","GV","VISITANTE","Fecha (Col)","Estadio"]
anch=[5,7,18,5,5,18,22,26]
wg.cell(row=1,column=1,value="🐔 Escribe los goles SOLO en las columnas amarillas (GL = goles local · GV = goles visita)").font=Bd
for col,(h,w) in enumerate(zip(heads,anch),1):
    style(wg.cell(row=2,column=col,value=h),H,INK); wg.column_dimensions[get_column_letter(col)].width=w
wg.freeze_panes="A3"
dvg=DataValidation(type="whole",operator="between",formula1="0",formula2="15",
                   errorTitle="¿Tantos goles?",error="Entre 0 y 15, parcero 😅"); wg.add_data_validation(dvg)
group_rows={}  # grupo -> (fila_inicio, fila_fin)
r=3
for g in sorted(grupos):
    ini=r
    for m in grupos[g]:
        vals=[m["MatchNumber"],g,nombre(m["HomeTeam"]),None,None,nombre(m["AwayTeam"]),fecha(m),m["Location"]]
        col_colombia = "Colombia" in (m["HomeTeam"],m["AwayTeam"])
        for col,v in enumerate(vals,1):
            c=wg.cell(row=r,column=col,value=v)
            style(c,N,("FFF3C4" if col_colombia else None),align=("left" if col in (3,6,7,8) else "center"))
        for col in (4,5):
            wg.cell(row=r,column=col).fill=PatternFill("solid",fgColor=YELLOW); dvg.add(wg.cell(row=r,column=col))
        r+=1
    group_rows[g]=(ini,r-1)
    r+=1  # fila en blanco entre grupos

# ===================== HOJA 3: TABLAS (auto-calculadas) =====================
wt=wb.create_sheet("TABLAS"); wt.sheet_properties.tabColor=BLUE
for col,w in zip("ABCDEFGHIJ",[5,20,5,5,5,5,6,6,6,6]): wt.column_dimensions[col].width=w
th=["POS","EQUIPO","PJ","G","E","P","GF","GC","DIF","PTS"]
GS="'FASE DE GRUPOS'"
row=1
wt.cell(row=row,column=1,value="📊 TABLAS DE POSICIONES — se llenan solas con los goles de la hoja anterior. Verde = clasifica 1º/2º.").font=Bd
row=2
for g in sorted(group_rows):
    r0,r1=group_rows[g]
    # título de grupo
    cab=wt.cell(row=row,column=1,value=f"GRUPO {g}"); cab.font=Font(name="Arial Black",size=12,color=INK)
    cab.fill=PatternFill("solid",fgColor=YELLOW);
    for col in range(1,11): wt.cell(row=row,column=col).fill=PatternFill("solid",fgColor=YELLOW); wt.cell(row=row,column=col).border=B
    row+=1
    for col,h in enumerate(th,1): style(wt.cell(row=row,column=col,value=h),H,INK)
    row+=1
    # 4 equipos (tomados de las filas del grupo: locales únicos + visitantes únicos)
    equipos=[]
    for rr in range(r0,r1+1):
        for cc in (3,6):
            v=wg.cell(row=rr,column=cc).value
            if v and v not in equipos: equipos.append(v)
    L=f"{GS}!$C${r0}:$C${r1}"; D=f"{GS}!$D${r0}:$D${r1}"; E=f"{GS}!$E${r0}:$E${r1}"; V=f"{GS}!$F${r0}:$F${r1}"
    primera_fila_equipos=row
    for eq in equipos:
        b=f'$B${row}'
        wt.cell(row=row,column=2,value=eq)
        # PJ
        wt.cell(row=row,column=3,value=f'=SUMPRODUCT(({L}={b})*({D}<>""))+SUMPRODUCT(({V}={b})*({E}<>""))')
        # G
        wt.cell(row=row,column=4,value=f'=SUMPRODUCT(({L}={b})*({D}>{E})*({D}<>""))+SUMPRODUCT(({V}={b})*({E}>{D})*({E}<>""))')
        # E
        wt.cell(row=row,column=5,value=f'=SUMPRODUCT(({L}={b})*({D}={E})*({D}<>""))+SUMPRODUCT(({V}={b})*({E}={D})*({E}<>""))')
        # P
        wt.cell(row=row,column=6,value=f'=$C${row}-$D${row}-$E${row}')
        # GF
        wt.cell(row=row,column=7,value=f'=SUMPRODUCT(({L}={b})*{D})+SUMPRODUCT(({V}={b})*{E})')
        # GC
        wt.cell(row=row,column=8,value=f'=SUMPRODUCT(({L}={b})*{E})+SUMPRODUCT(({V}={b})*{D})')
        # DIF
        wt.cell(row=row,column=9,value=f'=$G${row}-$H${row}')
        # PTS
        wt.cell(row=row,column=10,value=f'=3*$D${row}+$E${row}')
        for col in range(1,11):
            style(wt.cell(row=row,column=col),Bd if col in (2,10) else N, align=("left" if col==2 else "center"))
        row+=1
    # POS: 1 + nº de equipos con mejor llave (PTS, luego DIF, luego GF)
    pf=primera_fila_equipos; pl=row-1
    PTS=f"$J${pf}:$J${pl}"; DIF=f"$I${pf}:$I${pl}"; GF=f"$G${pf}:$G${pl}"
    for rr in range(pf,pl+1):
        f=(f'=1+SUMPRODUCT(({PTS}>$J${rr})*1)'
           f'+SUMPRODUCT(({PTS}=$J${rr})*({DIF}>$I${rr})*1)'
           f'+SUMPRODUCT(({PTS}=$J${rr})*({DIF}=$I${rr})*({GF}>$G${rr})*1)')
        cpos=wt.cell(row=rr,column=1,value=f); cpos.font=Font(name="Arial Black",size=11)
        # clasifican 1º y 2º → verde clarito (formato condicional simple: lo dejamos por fórmula de color manual no es trivial; resaltamos por valor con regla)
    row+=1

from openpyxl.formatting.rule import CellIsRule
# Regla: POS 1 o 2 en verde (sobre toda la columna A de TABLAS)
verde=PatternFill("solid",fgColor="C8E6C9")
wt.conditional_formatting.add(f"A2:A{row}", CellIsRule(operator="lessThanOrEqual",formula=["2"],fill=verde))

# ===================== HOJA 4: ELIMINATORIAS =====================
we=wb.create_sheet("ELIMINATORIAS"); we.sheet_properties.tabColor=RED
heads=["P#","FASE","LOCAL","","VISITANTE","Fecha (Col)","Estadio"]
anch=[5,10,22,4,22,22,26]
for col,(h,w) in enumerate(zip(heads,anch),1):
    style(we.cell(row=1,column=col,value=h),H,INK); we.column_dimensions[get_column_letter(col)].width=w
we.freeze_panes="A2"
r=2
for m in fixture:
    if m["Group"]: continue
    vals=[m["MatchNumber"],FASE[m["RoundNumber"]],nombre(m["HomeTeam"]),"vs",nombre(m["AwayTeam"]),fecha(m),m["Location"]]
    for col,v in enumerate(vals,1):
        style(we.cell(row=r,column=col,value=v),Bd if m["MatchNumber"]==104 else N, ("FFE08A" if m["MatchNumber"]==104 else None),
              align=("left" if col in (3,5,6,7) else "center"))
    r+=1

wb.save("public/fixture-mundial.xlsx")
print("OK: public/fixture-mundial.xlsx · grupos:",len(group_rows))
