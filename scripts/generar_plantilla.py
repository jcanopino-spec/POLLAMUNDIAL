# Genera public/plantilla-polla.xlsx: plantilla formulada de la Polla de Alameda
# para llenar pronósticos en Excel. Uso: python3 scripts/generar_plantilla.py
import json
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INK = "1B1714"; CREAM = "FBF4E6"; YELLOW = "FFC22E"; GREEN = "1B9150"; RED = "E1382F"; BLUE = "3447D6"; PAPER = "FFFFFF"

ES = {
    "Mexico": "México", "South Africa": "Sudáfrica", "Korea Republic": "Corea del Sur", "Czechia": "Chequia",
    "Canada": "Canadá", "Bosnia and Herzegovina": "Bosnia y Herzegovina", "Qatar": "Catar", "Switzerland": "Suiza",
    "Brazil": "Brasil", "Morocco": "Marruecos", "Haiti": "Haití", "Scotland": "Escocia",
    "USA": "Estados Unidos", "Paraguay": "Paraguay", "Australia": "Australia", "Türkiye": "Turquía",
    "Germany": "Alemania", "Curaçao": "Curazao", "Côte d'Ivoire": "Costa de Marfil", "Ecuador": "Ecuador",
    "Netherlands": "Países Bajos", "Japan": "Japón", "Sweden": "Suecia", "Tunisia": "Túnez",
    "Belgium": "Bélgica", "Egypt": "Egipto", "IR Iran": "Irán", "New Zealand": "Nueva Zelanda",
    "Spain": "España", "Cabo Verde": "Cabo Verde", "Saudi Arabia": "Arabia Saudita", "Uruguay": "Uruguay",
    "France": "Francia", "Senegal": "Senegal", "Norway": "Noruega", "Iraq": "Irak",
    "Argentina": "Argentina", "Algeria": "Argelia", "Austria": "Austria", "Jordan": "Jordania",
    "Colombia": "Colombia 🇨🇴", "Portugal": "Portugal", "Uzbekistan": "Uzbekistán", "Congo DR": "RD Congo",
    "England": "Inglaterra", "Croatia": "Croacia", "Ghana": "Ghana", "Panama": "Panamá",
}

FASE = {1: "Grupos · F1", 2: "Grupos · F2", 3: "Grupos · F3", 4: "16avos", 5: "Octavos", 6: "Cuartos", 7: "Semis", 8: "Final"}


def nombre(eq: str) -> str:
    if eq in ES:
        return ES[eq]
    import re
    m = re.match(r"^([123])([A-L])$", eq)
    if m:
        return f"{m.group(1)}º Grupo {m.group(2)}"
    m = re.match(r"^3([A-L]{2,})$", eq)
    if m:
        return "3º de " + "/".join(m.group(1))
    m = re.match(r"^W(\d+)$", eq)
    if m:
        return f"Ganador P{m.group(1)}"
    return "Por definir"


def mult(match_id: int, rnd: int) -> int:
    if match_id == 104:
        return 6
    return {1: 1, 2: 1, 3: 1, 4: 2, 5: 3, 6: 4, 7: 5, 8: 5}[rnd]


fixture = sorted(json.load(open("data/fixture-raw.json")), key=lambda m: (m["DateUtc"], m["MatchNumber"]))
wb = Workbook()

thin = Side(style="thin", color=INK)
borde = Border(left=thin, right=thin, top=thin, bottom=thin)
f_titulo = Font(name="Arial Black", size=16, color=INK)
f_header = Font(bold=True, color=PAPER, size=10)
f_normal = Font(size=10)

# ============ HOJA 1: LÉEME 🐔 ============
ws = wb.active
ws.title = "LÉEME 🐔"
ws.sheet_properties.tabColor = YELLOW
ws.column_dimensions["A"].width = 100
filas = [
    ("🐔⚽ LA POLLA MUNDIALISTA DE ALAMEDA — VERSIÓN EXCEL ⚽🐔", f_titulo, YELLOW),
    ("", None, None),
    ("¿Te da miedo la app? ¿Eres de los que todavía imprime los correos? Tranquilo parcero, esta plantilla es para ti 😂", f_normal, None),
    ("", None, None),
    ("CÓMO SE LLENA (hasta Doña Julia puede):", Font(bold=True, size=12), CREAM),
    ("1. Ve a la hoja 'PRONÓSTICOS' y escribe tus goles en las columnas amarillas (MI LOCAL y MI VISITA).", f_normal, None),
    ("2. En la hoja 'APUESTAS GRANDES' elige tus 2 finalistas y tu campeón (hay lista desplegable, no inventes equipos).", f_normal, None),
    ("3. Cuando haya resultados reales, se escriben en las columnas verdes y EL EXCEL CALCULA TUS PUNTOS SOLO. Magia pura 🧙", f_normal, None),
    ("4. Envíale el archivo lleno al admin ANTES del jueves 11 de junio 2:00 PM… él lo pasa a la app (y te cobra el favor 😏).", f_normal, None),
    ("", None, None),
    ("EL PUNTAJE (apréndetelo):", Font(bold=True, size=12), CREAM),
    ("🎯 Marcador exacto = 5 pts · ✔️ Acertar ganador o empate = 3 pts", f_normal, None),
    ("📈 Y entre más avanza el Mundial, más vale: 16avos ×2 · Octavos ×3 · Cuartos ×4 · Semis ×5 · LA FINAL ×6 (exacto = 30 puntazos)", f_normal, None),
    ("⭐ Finalista acertado = 15 pts c/u · 👑 Campeón = 30 pts", f_normal, None),
    ("⛔ En la app los pronósticos se cierran cuando pita el árbitro. En Excel confiamos en tu honor… y en el admin 🕵️", f_normal, None),
    ("🥅 En eliminatorias cuenta el marcador final con prórroga (los penales solo dicen quién pasa).", f_normal, None),
    ("", None, None),
    ("🐷 EL CASTIGO: el último de la polla pone el guaro y el cerdo de la porcícola. Quedas avisado.", Font(bold=True, size=11, color=RED), None),
    ("💰 Esto es pa' gozarnos el Mundial, reírnos y recoger billetico pa' los fondos de la natillera.", f_normal, None),
    ("", None, None),
    ("La app oficial (más fácil que este Excel, de verdad): pollamundialnatillera.vercel.app", Font(bold=True, color=BLUE.replace('#',''), size=11), None),
    ("Diseñada por Jcanopino · CEO INPLUX SAS — prohibido copiarla: la gallina tiene abogados 🐔⚖️", Font(italic=True, size=9), None),
]
for i, (txt, fnt, fill) in enumerate(filas, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    if fnt:
        c.font = fnt
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(wrap_text=True, vertical="center")

# ============ HOJA 2: PRONÓSTICOS ============
ws = wb.create_sheet("PRONÓSTICOS")
ws.sheet_properties.tabColor = GREEN
headers = ["P#", "FASE", "FECHA (HORA COLOMBIA)", "ESTADIO", "LOCAL", "MI LOCAL ✍️", "MI VISITA ✍️", "VISITANTE", "VALE ×", "REAL LOCAL", "REAL VISITA", "MIS PUNTOS 🤑"]
anchos = [5, 12, 22, 24, 22, 11, 11, 22, 7, 11, 11, 13]
ws.cell(row=1, column=1, value="🐔 TUS PRONÓSTICOS — llena SOLO las columnas amarillas. Los puntos se calculan solos cuando haya resultados reales.").font = Font(bold=True, size=11)
ws.cell(row=2, column=11, value="TOTAL →").font = Font(bold=True, size=12)
ws.cell(row=2, column=12, value=f"=SUM(L4:L{3 + len(fixture)})").font = Font(bold=True, size=14, color=GREEN)
ws.cell(row=2, column=12).fill = PatternFill("solid", fgColor=YELLOW)
ws.cell(row=2, column=12).border = borde
for col, (h, w) in enumerate(zip(headers, anchos), start=1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = f_header
    c.fill = PatternFill("solid", fgColor=INK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = borde
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A4"

dv_goles = DataValidation(type="whole", operator="between", formula1="0", formula2="15",
                          errorTitle="¿Tantos goles, parcero?", error="Entre 0 y 15… esto es fútbol, no tejo 😂")
ws.add_data_validation(dv_goles)

bogota = ZoneInfo("America/Bogota")
DIAS = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]
MESES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]
for i, m in enumerate(fixture):
    r = 4 + i
    dt = datetime.fromisoformat(m["DateUtc"].replace(" ", "T").replace("Z", "+00:00")).astimezone(bogota)
    fecha = f"{DIAS[dt.weekday()]} {dt.day} {MESES[dt.month - 1]} · {dt.strftime('%I:%M %p').lstrip('0')}"
    rnd = m["RoundNumber"]
    vale = mult(m["MatchNumber"], rnd)
    vals = [m["MatchNumber"], FASE[rnd], fecha, m["Location"],
            nombre(m["HomeTeam"]), None, None, nombre(m["AwayTeam"]), vale, None, None]
    for col, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=col, value=v)
        c.font = f_normal
        c.border = borde
        c.alignment = Alignment(horizontal="center" if col not in (4, 5, 8) else "left", vertical="center")
        if rnd <= 3 and "Colombia" in (m["HomeTeam"], m["AwayTeam"]):
            c.fill = PatternFill("solid", fgColor="FFF3C4")
    for col in (6, 7):  # mis goles → amarillo
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=YELLOW)
        dv_goles.add(ws.cell(row=r, column=col))
    for col in (10, 11):  # resultado real → verde clarito
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="DDF0E4")
    # puntos: exacto 5×mult, resultado 3×mult, nada 0 — con humor en el cero
    f = (f'=IF(OR(F{r}="",G{r}="",J{r}="",K{r}=""),"",'
         f'IF(AND(F{r}=J{r},G{r}=K{r}),5*I{r},'
         f'IF(SIGN(F{r}-G{r})=SIGN(J{r}-K{r}),3*I{r},0)))')
    c = ws.cell(row=r, column=12, value=f)
    c.font = Font(bold=True, size=10)
    c.border = borde
    c.alignment = Alignment(horizontal="center")

# ============ HOJA 3: APUESTAS GRANDES ============
ws = wb.create_sheet("APUESTAS GRANDES")
ws.sheet_properties.tabColor = RED
for col, w in zip("ABCD", [30, 26, 26, 16]):
    ws.column_dimensions[col].width = w
ws["A1"] = "💰 LAS APUESTAS GRANDES — elige de la lista, no inventes selecciones"
ws["A1"].font = Font(bold=True, size=12)
rows = [
    ("TU FINALISTA #1 (15 pts si llega)", "B3"),
    ("TU FINALISTA #2 (15 pts si llega)", "B4"),
    ("TU CAMPEÓN 👑 (30 pts · debe ser uno de tus 2 finalistas)", "B5"),
]
equipos = sorted(ES.values())
hidden = wb.create_sheet("_equipos")
hidden.sheet_state = "hidden"
for i, eq in enumerate(equipos, start=1):
    hidden.cell(row=i, column=1, value=eq)
dv_eq = DataValidation(type="list", formula1=f"=_equipos!$A$1:$A${len(equipos)}",
                       errorTitle="Ese equipo no existe", error="De la lista, parcero. Vaticano FC no clasificó 😂")
ws.add_data_validation(dv_eq)
for i, (label, cell) in enumerate(rows, start=3):
    ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
    c = ws[cell]
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.border = borde
    dv_eq.add(c)
ws["A7"] = "— Cuando se sepan los de verdad (los llena el admin) —"
ws["A7"].font = Font(italic=True, size=9)
ws["A8"], ws["A9"], ws["A10"] = "FINALISTA REAL #1", "FINALISTA REAL #2", "CAMPEÓN REAL"
for r in (8, 9, 10):
    ws.cell(row=r, column=1).font = Font(bold=True, size=10)
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="DDF0E4")
    ws.cell(row=r, column=2).border = borde
ws["A12"] = "TUS PUNTOS DE BONO →"
ws["A12"].font = Font(bold=True, size=12)
ws["B12"] = ('=IF(OR($B$8="",$B$9=""),"",'
             'SUMPRODUCT(--(COUNTIF($B$8:$B$9,$B$3:$B$4)>0))*15 + IF($B$5=$B$10,30,0))')
ws["B12"].font = Font(bold=True, size=14, color=GREEN)
ws["B12"].fill = PatternFill("solid", fgColor=YELLOW)
ws["B12"].border = borde
ws["A14"] = "💡 Pro tip: antes de apostar, prueba el SIMULADOR de la app 🔮 — verifica que tu final SÍ exista y no se cruce en octavos."
ws["A14"].font = Font(italic=True, size=9)

wb.save("public/plantilla-polla.xlsx")
print("OK: public/plantilla-polla.xlsx")
